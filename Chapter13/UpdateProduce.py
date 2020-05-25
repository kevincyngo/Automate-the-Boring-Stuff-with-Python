# Loops over all the rows
# If the row is for garlic, celery, or lemons, changes the price

# This means your code will need to do the following:
# Open the spreadsheet file.
# For each row, check whether the value in column A is Celery, Garlic, or Lemon.
# If it is, update the price in column B.
# Save the spreadsheet to a new file (so that you don’t lose the old spreadsheet, just in case).

import openpyxl

#Set up Data Structure with update information
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
