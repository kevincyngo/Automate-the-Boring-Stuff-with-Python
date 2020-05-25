# Import the openpyxl module.
# Call the openpyxl.load_workbook() function.
# Get a Workbook object.
# Use the active or sheetnames attributes.
# Get a Worksheet object.
# Use indexing or the cell() sheet method with row and column keyword arguments.
# Get a Cell object.
# Read the Cell object’s value attribute.

import openpyxl, pprint

print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
countyData = {}

#Fill in countyData with each county's population and tracts
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)

#Open new text file and write the contents of countyData to it
print('Writing results...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')
